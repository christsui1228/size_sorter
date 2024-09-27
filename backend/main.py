import pandas as pd
import os
import sys
import logging
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import string

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 定义尺码排序顺序
SIZE_ORDER = ['100','110','120','130','140','150','XS','S', 'M', 'L', 'XL', '2XL', '3XL', '4XL', '5XL','6XL','7XL','8XL','9XL','10XL']

def get_input_path():
    return input("请输入要处理的Excel文件的完整路径：").strip()

def get_output_dir(input_path):
    output_dir = input("请输入保存结果的目录路径（按回车键使用默认输出目录）：").strip()
    return output_dir if output_dir else os.path.dirname(input_path)

def get_rows_per_column():
    while True:
        try:
            rows = int(input("请输入每列显示的行数（不包括标题行）："))
            if rows > 0:
                return rows
            else:
                print("请输入一个正整数。")
        except ValueError:
            print("无效输入，请输入一个整数。")

def clean_and_order_size(size):
    size = str(size).upper().strip()
    if size in SIZE_ORDER:
        return SIZE_ORDER.index(size)
    elif size.endswith('XL'):
        x_count = size.count('X')
        if x_count > 1:
            return SIZE_ORDER.index('XL') + x_count
    return len(SIZE_ORDER)  # 未知尺码放到最后

def process_excel(input_path, output_dir, rows_per_column):
    try:
        # 读取Excel文件
        df = pd.read_excel(input_path)
        logging.info(f"成功读取Excel文件，共 {len(df)} 行")
        
        # 确保列名正确
        if len(df.columns) >= 2:
            df = df.iloc[:, :2]  # 只取前两列
            df.columns = ['姓名', '尺码']
        else:
            raise ValueError("Excel文件应至少包含两列数据（姓名和尺码）")

        # 数据清洗和排序
        df['尺码排序'] = df['尺码'].apply(clean_and_order_size)
        df['姓名长度'] = df['姓名'].str.len()
        df = df.sort_values(['尺码排序', '姓名长度', '姓名'])
        df = df.drop(['尺码排序', '姓名长度'], axis=1)

        # 添加序号列
        df['序号'] = range(1, len(df) + 1)
        df = df[['序号', '姓名', '尺码']]

        # 创建新的Excel文件
        source_filename = os.path.splitext(os.path.basename(input_path))[0]
        output_filename = f'{source_filename}_sorted_formatted.xlsx'
        output_path = os.path.join(output_dir, output_filename)
        
        # 使用openpyxl创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "排序后数据"

        # 生成动态列组
        all_columns = list(string.ascii_uppercase) + [f'A{c}' for c in string.ascii_uppercase]
        column_groups = [all_columns[i:i+3] for i in range(0, len(all_columns), 3)]

        # 写入标题和数据并设置格式
        headers = ['序号', '姓名', '尺码']
        group_index = 0
        row_index = 0
        while row_index < len(df):
            column_group = column_groups[group_index]
            # 写入标题
            for col, header in zip(column_group, headers):
                cell = ws[f'{col}1']
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # 写入数据
            end_row = min(row_index + rows_per_column, len(df))
            for i, row in enumerate(df.iloc[row_index:end_row].values, start=2):
                for j, value in enumerate(row):
                    column = column_group[j]
                    cell = ws[f'{column}{i}']
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center' if j != 1 else 'left')

            row_index = end_row
            group_index += 1

        # 设置列宽
        for column_group in column_groups[:group_index]:
            for column in column_group:
                ws.column_dimensions[column].width = 15

        # 保存文件
        wb.save(output_path)
        logging.info(f"文件已成功保存: {output_path}")

    except Exception as e:
        logging.error(f"处理过程中出现错误: {str(e)}")
        raise

def main():
    start_time = time.time()  # 记录开始时间

    logging.info(f"Python version: {sys.version}")
    logging.info(f"Pandas version: {pd.__version__}")

    input_path = get_input_path()
    output_dir = get_output_dir(input_path)
    rows_per_column = get_rows_per_column()
    os.makedirs(output_dir, exist_ok=True)

    try:
        process_excel(input_path, output_dir, rows_per_column)
        logging.info("处理完成。请检查输出文件。")
    except Exception as e:
        logging.error(f"程序执行失败: {str(e)}")

    end_time = time.time()  # 记录结束时间
    execution_time = end_time - start_time  # 计算执行时间

    logging.info(f"程序执行时间: {execution_time:.2f} 秒")  # 显示程序执行时间

if __name__ == "__main__":
    main()